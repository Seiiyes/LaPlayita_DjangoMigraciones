from django.shortcuts import render
from django.db.models import Sum, Count, Avg, F, Q, Max, Min
from django.db.models.functions import TruncDate, TruncHour
from pos.models import Venta, VentaDetalle, Pago
from inventory.models import Producto, Lote, MovimientoInventario
from clients.models import Cliente, PuntosFidelizacion
from django.utils.dateparse import parse_date
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from datetime import date, timedelta
from decimal import Decimal
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def panel_reportes(request):
    return render(request, 'reportes/panel_reportes.html', {})

def reporte_ventas(request):
    """Reporte mejorado de ventas con análisis detallado"""
    # Obtener parámetros de fecha
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    metodo_pago = request.GET.get('metodo_pago', '')
    canal_venta = request.GET.get('canal_venta', '')

    # Filtrar ventas
    ventas = Venta.objects.select_related('cliente', 'usuario').all()
    
    if fecha_inicio:
        ventas = ventas.filter(fecha_venta__date__gte=parse_date(fecha_inicio))
    if fecha_fin:
        ventas = ventas.filter(fecha_venta__date__lte=parse_date(fecha_fin))
    if metodo_pago:
        ventas = ventas.filter(pago__metodo_pago=metodo_pago)
    if canal_venta:
        ventas = ventas.filter(canal_venta=canal_venta)

    # Ventas del día
    hoy = timezone.now().date()
    ventas_hoy = Venta.objects.filter(fecha_venta__date=hoy)

    # KPIs principales
    total_ventas = ventas.aggregate(total=Sum('total_venta'))['total'] or Decimal('0')
    total_ventas_hoy = ventas_hoy.aggregate(total=Sum('total_venta'))['total'] or Decimal('0')
    cantidad_ventas = ventas.count()
    cantidad_ventas_hoy = ventas_hoy.count()
    ticket_promedio = total_ventas / cantidad_ventas if cantidad_ventas > 0 else Decimal('0')

    # Ventas por método de pago
    ventas_por_metodo = Pago.objects.filter(venta__in=ventas).values('metodo_pago').annotate(
        total=Sum('monto'),
        cantidad=Count('id')
    ).order_by('-total')

    # Ventas por canal
    ventas_por_canal = ventas.values('canal_venta').annotate(
        total=Sum('total_venta'),
        cantidad=Count('id')
    ).order_by('-total')

    # Productos más vendidos
    productos_mas_vendidos = VentaDetalle.objects.filter(venta__in=ventas).values(
        'producto__nombre'
    ).annotate(
        cantidad_total=Sum('cantidad'),
        ingresos=Sum('subtotal')
    ).order_by('-cantidad_total')[:10]

    # Ventas por hora (para identificar horas pico)
    ventas_por_hora = ventas.annotate(
        hora=TruncHour('fecha_venta')
    ).values('hora').annotate(
        total=Sum('total_venta'),
        cantidad=Count('id')
    ).order_by('hora')

    # Exportar a Excel
    if 'descargar' in request.GET:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="reporte_ventas_detallado.xlsx"'

        wb = Workbook()
        
        # Estilos
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=14, color="2F5496")
        currency_format = '#,##0.00'
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Hoja 1: Resumen
        ws1 = wb.active
        ws1.title = "Resumen"
        
        # Título
        ws1['A1'] = 'REPORTE DE VENTAS - LA PLAYITA'
        ws1['A1'].font = Font(bold=True, size=16, color="2F5496")
        ws1.merge_cells('A1:B1')
        ws1.append([])
        
        # KPIs
        ws1.append(['Total Ventas:', float(total_ventas)])
        ws1['B3'].number_format = currency_format
        ws1['A3'].font = Font(bold=True)
        
        ws1.append(['Cantidad Ventas:', cantidad_ventas])
        ws1['A4'].font = Font(bold=True)
        
        ws1.append(['Ticket Promedio:', float(ticket_promedio)])
        ws1['B5'].number_format = currency_format
        ws1['A5'].font = Font(bold=True)
        
        ws1.append([])
        
        # Tabla de métodos de pago
        ws1.append(['Método de Pago', 'Total', 'Cantidad'])
        for cell in ws1[7]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
        
        for metodo in ventas_por_metodo:
            ws1.append([metodo['metodo_pago'].title(), float(metodo['total']), metodo['cantidad']])
            ws1.cell(ws1.max_row, 2).number_format = currency_format
        
        # Ajustar anchos
        ws1.column_dimensions['A'].width = 20
        ws1.column_dimensions['B'].width = 15

        # Hoja 2: Detalle de ventas
        ws2 = wb.create_sheet("Detalle Ventas")
        ws2.append(['ID', 'Fecha', 'Cliente', 'Canal', 'Método Pago', 'Total'])
        
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
        
        for venta in ventas:
            pago = venta.pago_set.first()
            ws2.append([
                venta.id,
                venta.fecha_venta.strftime('%d/%m/%Y %H:%M'),
                f"{venta.cliente.nombres} {venta.cliente.apellidos}",
                venta.canal_venta.title(),
                pago.metodo_pago.title() if pago else 'N/A',
                float(venta.total_venta)
            ])
            ws2.cell(ws2.max_row, 6).number_format = currency_format
        
        # Ajustar anchos
        ws2.column_dimensions['A'].width = 8
        ws2.column_dimensions['B'].width = 18
        ws2.column_dimensions['C'].width = 30
        ws2.column_dimensions['D'].width = 12
        ws2.column_dimensions['E'].width = 15
        ws2.column_dimensions['F'].width = 12

        # Hoja 3: Productos más vendidos
        ws3 = wb.create_sheet("Top Productos")
        ws3.append(['Producto', 'Cantidad Vendida', 'Ingresos'])
        
        for cell in ws3[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
        
        for prod in productos_mas_vendidos:
            ws3.append([prod['producto__nombre'], prod['cantidad_total'], float(prod['ingresos'])])
            ws3.cell(ws3.max_row, 3).number_format = currency_format
        
        # Ajustar anchos
        ws3.column_dimensions['A'].width = 40
        ws3.column_dimensions['B'].width = 18
        ws3.column_dimensions['C'].width = 15

        wb.save(response)
        return response

    return render(request, 'reportes/reporte_ventas.html', {
        'ventas': ventas[:50],  # Limitar a 50 para performance
        'total_ventas': total_ventas,
        'total_ventas_hoy': total_ventas_hoy,
        'cantidad_ventas': cantidad_ventas,
        'cantidad_ventas_hoy': cantidad_ventas_hoy,
        'ticket_promedio': ticket_promedio,
        'ventas_por_metodo': ventas_por_metodo,
        'ventas_por_canal': ventas_por_canal,
        'productos_mas_vendidos': productos_mas_vendidos,
        'ventas_por_hora': ventas_por_hora,
    })


def reporte_inventario(request):
    """Reporte completo de inventario"""
    # Productos con stock bajo
    productos_stock_bajo = Producto.objects.filter(
        stock_actual__lte=F('stock_minimo')
    ).order_by('stock_actual')

    # Productos próximos a vencer (30 días)
    fecha_limite = timezone.now().date() + timedelta(days=30)
    lotes_por_vencer = Lote.objects.filter(
        fecha_caducidad__lte=fecha_limite,
        cantidad_disponible__gt=0
    ).select_related('producto').order_by('fecha_caducidad')

    # Valor total del inventario
    valor_inventario = Producto.objects.aggregate(
        valor_total=Sum(F('stock_actual') * F('precio_unitario'))
    )['valor_total'] or Decimal('0')

    # Productos sin movimiento (últimos 30 días)
    hace_30_dias = timezone.now() - timedelta(days=30)
    productos_con_movimiento = MovimientoInventario.objects.filter(
        fecha_movimiento__gte=hace_30_dias
    ).values_list('producto_id', flat=True).distinct()
    
    productos_sin_movimiento = Producto.objects.exclude(
        id__in=productos_con_movimiento
    ).filter(stock_actual__gt=0)

    # Top productos por rotación
    productos_mas_vendidos = VentaDetalle.objects.filter(
        venta__fecha_venta__gte=hace_30_dias
    ).values('producto__nombre').annotate(
        cantidad_vendida=Sum('cantidad'),
        ingresos=Sum('subtotal')
    ).order_by('-cantidad_vendida')[:10]

    # Movimientos recientes
    movimientos_recientes = MovimientoInventario.objects.select_related(
        'producto', 'lote'
    ).order_by('-fecha_movimiento')[:20]

    # Exportar a Excel
    if 'descargar' in request.GET:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="reporte_inventario.xlsx"'

        wb = Workbook()
        
        # Estilos
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        danger_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        currency_format = '#,##0.00'
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Hoja 1: Stock Bajo
        ws1 = wb.active
        ws1.title = "Stock Bajo"
        ws1.append(['Producto', 'Stock Actual', 'Stock Mínimo', 'Precio', 'Valor Total'])
        
        for cell in ws1[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
        
        for prod in productos_stock_bajo:
            row = [
                prod.nombre,
                prod.stock_actual,
                prod.stock_minimo,
                float(prod.precio_unitario),
                float(prod.stock_actual * prod.precio_unitario)
            ]
            ws1.append(row)
            
            # Colorear según criticidad
            if prod.stock_actual == 0:
                ws1.cell(ws1.max_row, 2).fill = danger_fill
            else:
                ws1.cell(ws1.max_row, 2).fill = warning_fill
            
            ws1.cell(ws1.max_row, 4).number_format = currency_format
            ws1.cell(ws1.max_row, 5).number_format = currency_format
        
        ws1.column_dimensions['A'].width = 40
        ws1.column_dimensions['B'].width = 12
        ws1.column_dimensions['C'].width = 12
        ws1.column_dimensions['D'].width = 12
        ws1.column_dimensions['E'].width = 15

        # Hoja 2: Próximos a Vencer
        ws2 = wb.create_sheet("Por Vencer")
        ws2.append(['Producto', 'Lote', 'Cantidad', 'Fecha Caducidad', 'Días Restantes'])
        
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
        
        for lote in lotes_por_vencer:
            dias_restantes = (lote.fecha_caducidad - timezone.now().date()).days
            ws2.append([
                lote.producto.nombre,
                lote.numero_lote,
                lote.cantidad_disponible,
                lote.fecha_caducidad.strftime('%d/%m/%Y'),
                dias_restantes
            ])
            
            # Colorear según urgencia
            if dias_restantes <= 7:
                ws2.cell(ws2.max_row, 5).fill = danger_fill
            else:
                ws2.cell(ws2.max_row, 5).fill = warning_fill
        
        ws2.column_dimensions['A'].width = 40
        ws2.column_dimensions['B'].width = 20
        ws2.column_dimensions['C'].width = 12
        ws2.column_dimensions['D'].width = 18
        ws2.column_dimensions['E'].width = 15

        # Hoja 3: Sin Movimiento
        ws3 = wb.create_sheet("Sin Movimiento")
        ws3.append(['Producto', 'Stock', 'Valor Inmovilizado'])
        
        for cell in ws3[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
        
        for prod in productos_sin_movimiento:
            ws3.append([
                prod.nombre,
                prod.stock_actual,
                float(prod.stock_actual * prod.precio_unitario)
            ])
            ws3.cell(ws3.max_row, 3).number_format = currency_format
        
        ws3.column_dimensions['A'].width = 40
        ws3.column_dimensions['B'].width = 12
        ws3.column_dimensions['C'].width = 18

        wb.save(response)
        return response

    return render(request, 'reportes/reporte_inventario.html', {
        'productos_stock_bajo': productos_stock_bajo,
        'lotes_por_vencer': lotes_por_vencer,
        'valor_inventario': valor_inventario,
        'productos_sin_movimiento': productos_sin_movimiento,
        'productos_mas_vendidos': productos_mas_vendidos,
        'movimientos_recientes': movimientos_recientes,
    })


def reporte_clientes(request):
    """Reporte completo de clientes"""
    # Excluir "Consumidor Final"
    clientes = Cliente.objects.exclude(id=1)

    # Top clientes por compras
    top_clientes = clientes.annotate(
        total_compras=Sum('venta__total_venta'),
        cantidad_compras=Count('venta'),
        ultima_compra=Max('venta__fecha_venta')
    ).filter(total_compras__isnull=False).order_by('-total_compras')[:20]

    # Clientes con más puntos (usar el campo existente del modelo)
    clientes_puntos = clientes.filter(
        puntos_totales__gt=0
    ).order_by('-puntos_totales')[:10]

    # Clientes nuevos (últimos 30 días) - basado en primera compra
    hace_30_dias = timezone.now() - timedelta(days=30)
    clientes_nuevos = clientes.annotate(
        primera_compra=Min('venta__fecha_venta')
    ).filter(
        primera_compra__gte=hace_30_dias
    ).order_by('-primera_compra')

    # Clientes inactivos (sin compras en 60 días)
    hace_60_dias = timezone.now() - timedelta(days=60)
    clientes_con_compras_recientes = Venta.objects.filter(
        fecha_venta__gte=hace_60_dias
    ).values_list('cliente_id', flat=True).distinct()
    
    clientes_inactivos = clientes.exclude(
        id__in=clientes_con_compras_recientes
    ).annotate(
        ultima_compra=Max('venta__fecha_venta')
    ).filter(ultima_compra__isnull=False).order_by('ultima_compra')

    # Estadísticas generales
    total_clientes = clientes.count()
    clientes_activos = clientes.filter(id__in=clientes_con_compras_recientes).count()
    ticket_promedio_general = Venta.objects.exclude(cliente_id=1).aggregate(
        promedio=Avg('total_venta')
    )['promedio'] or Decimal('0')

    # Exportar a Excel
    if 'descargar' in request.GET:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="reporte_clientes.xlsx"'

        wb = Workbook()
        
        # Estilos
        header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        gold_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
        danger_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        currency_format = '#,##0.00'
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Hoja 1: Top Clientes
        ws1 = wb.active
        ws1.title = "Top Clientes"
        ws1.append(['Cliente', 'Total Compras', 'Cantidad Compras', 'Ticket Promedio', 'Última Compra'])
        
        for cell in ws1[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
        
        for idx, cliente in enumerate(top_clientes, 1):
            ticket_prom = cliente.total_compras / cliente.cantidad_compras if cliente.cantidad_compras > 0 else 0
            ws1.append([
                f"{cliente.nombres} {cliente.apellidos}",
                float(cliente.total_compras),
                cliente.cantidad_compras,
                float(ticket_prom),
                cliente.ultima_compra.strftime('%d/%m/%Y') if cliente.ultima_compra else 'N/A'
            ])
            
            # Destacar top 3
            if idx <= 3:
                ws1.cell(ws1.max_row, 1).fill = gold_fill
            
            ws1.cell(ws1.max_row, 2).number_format = currency_format
            ws1.cell(ws1.max_row, 4).number_format = currency_format
        
        ws1.column_dimensions['A'].width = 35
        ws1.column_dimensions['B'].width = 15
        ws1.column_dimensions['C'].width = 18
        ws1.column_dimensions['D'].width = 18
        ws1.column_dimensions['E'].width = 15

        # Hoja 2: Clientes con Puntos
        ws2 = wb.create_sheet("Puntos Fidelización")
        ws2.append(['Cliente', 'Puntos Totales', 'Teléfono', 'Correo'])
        
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
        
        for cliente in clientes_puntos:
            ws2.append([
                f"{cliente.nombres} {cliente.apellidos}",
                float(cliente.puntos_totales),
                cliente.telefono,
                cliente.correo
            ])
            ws2.cell(ws2.max_row, 2).number_format = '#,##0'
        
        ws2.column_dimensions['A'].width = 35
        ws2.column_dimensions['B'].width = 15
        ws2.column_dimensions['C'].width = 15
        ws2.column_dimensions['D'].width = 30

        # Hoja 3: Clientes Inactivos
        ws3 = wb.create_sheet("Clientes Inactivos")
        ws3.append(['Cliente', 'Última Compra', 'Días Inactivo', 'Teléfono'])
        
        for cell in ws3[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
        
        for cliente in clientes_inactivos:
            dias_inactivo = (timezone.now().date() - cliente.ultima_compra.date()).days if cliente.ultima_compra else 0
            ws3.append([
                f"{cliente.nombres} {cliente.apellidos}",
                cliente.ultima_compra.strftime('%d/%m/%Y') if cliente.ultima_compra else 'N/A',
                dias_inactivo,
                cliente.telefono
            ])
            
            # Colorear según días inactivos
            if dias_inactivo > 90:
                ws3.cell(ws3.max_row, 3).fill = danger_fill
        
        ws3.column_dimensions['A'].width = 35
        ws3.column_dimensions['B'].width = 15
        ws3.column_dimensions['C'].width = 15
        ws3.column_dimensions['D'].width = 15

        wb.save(response)
        return response

    return render(request, 'reportes/reporte_clientes.html', {
        'top_clientes': top_clientes,
        'clientes_puntos': clientes_puntos,
        'clientes_nuevos': clientes_nuevos,
        'clientes_inactivos': clientes_inactivos,
        'total_clientes': total_clientes,
        'clientes_activos': clientes_activos,
        'ticket_promedio_general': ticket_promedio_general,
    })
