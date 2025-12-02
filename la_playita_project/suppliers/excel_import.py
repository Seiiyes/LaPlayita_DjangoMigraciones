"""
Importación masiva de reabastecimientos desde Excel
"""
from io import BytesIO
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from inventory.models import Producto
import logging

logger = logging.getLogger(__name__)


def generate_template_excel():
    """
    Genera una plantilla Excel para carga masiva de reabastecimientos
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Plantilla Reabastecimiento"
    
    # Estilos
    header_fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    instruction_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    instruction_font = Font(size=10, color="856404")
    example_fill = PatternFill(start_color="E7F1FF", end_color="E7F1FF", fill_type="solid")
    
    border = Border(
        left=Side(style='thin', color='DEE2E6'),
        right=Side(style='thin', color='DEE2E6'),
        top=Side(style='thin', color='DEE2E6'),
        bottom=Side(style='thin', color='DEE2E6')
    )
    
    # Título
    ws.merge_cells('A1:E1')
    ws['A1'] = "Plantilla de Carga Masiva - Reabastecimiento"
    ws['A1'].font = Font(bold=True, size=16, color="0D6EFD")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # Instrucciones
    ws.merge_cells('A2:E2')
    ws['A2'] = "INSTRUCCIONES: Complete los datos a partir de la fila 5. No modifique los encabezados."
    ws['A2'].font = instruction_font
    ws['A2'].fill = instruction_fill
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[2].height = 30
    
    ws.merge_cells('A3:E3')
    ws['A3'] = "Puede usar el ID o nombre del producto. La fecha debe estar en formato DD/MM/YYYY"
    ws['A3'].font = instruction_font
    ws['A3'].fill = instruction_fill
    ws['A3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[3].height = 25
    
    # Headers
    headers = ['ID Producto', 'Nombre Producto', 'Cantidad', 'Costo Unitario', 'Fecha Caducidad']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Ejemplo
    example_data = [
        ['1', 'Cerveza Corona', 100, 5000, '31/12/2025'],
        ['', 'Cerveza Aguila', 50, 4500, '30/11/2025'],
        ['3', '', 75, 3000, '15/10/2025'],
    ]
    
    for row_idx, row_data in enumerate(example_data, start=5):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = example_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center' if col_idx > 2 else 'left', vertical='center')
            
            # Formato de número para costo
            if col_idx == 4:
                cell.number_format = '#,##0'
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 18
    
    # Agregar nota al final
    ws.merge_cells('A8:E8')
    ws['A8'] = "NOTA: Los datos de ejemplo (filas 5-7) son solo de referencia. Elimínelos antes de cargar su archivo."
    ws['A8'].font = Font(size=9, italic=True, color="6C757D")
    ws['A8'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Guardar en buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer


def parse_reabastecimiento_excel(file):
    """
    Parsea un archivo Excel y extrae los datos de productos para reabastecimiento
    
    Returns:
        dict: {
            'success': bool,
            'data': list of dicts con productos,
            'errors': list of error messages,
            'warnings': list of warning messages
        }
    """
    result = {
        'success': False,
        'data': [],
        'errors': [],
        'warnings': []
    }
    
    try:
        # Cargar workbook
        wb = load_workbook(file, data_only=True)
        ws = wb.active
        
        # Validar que tenga datos
        if ws.max_row < 5:
            result['errors'].append('El archivo no contiene datos. Debe tener al menos una fila de datos después de los encabezados.')
            return result
        
        # Procesar filas (empezar desde la fila 5, después de headers y ejemplos)
        productos_procesados = []
        
        for row_idx in range(5, ws.max_row + 1):
            # Leer valores
            producto_id = ws.cell(row=row_idx, column=1).value
            nombre = ws.cell(row=row_idx, column=2).value
            cantidad = ws.cell(row=row_idx, column=3).value
            costo = ws.cell(row=row_idx, column=4).value
            fecha_cad = ws.cell(row=row_idx, column=5).value
            
            # Saltar filas vacías
            if not producto_id and not nombre and not cantidad:
                continue
            
            # Validar que tenga al menos ID o nombre
            if not producto_id and not nombre:
                result['warnings'].append(f'Fila {row_idx}: Se omitió porque no tiene ID ni nombre de producto.')
                continue
            
            # Validar cantidad
            if not cantidad or cantidad <= 0:
                result['errors'].append(f'Fila {row_idx}: La cantidad debe ser mayor a 0.')
                continue
            
            # Validar costo
            if not costo or costo <= 0:
                result['errors'].append(f'Fila {row_idx}: El costo unitario debe ser mayor a 0.')
                continue
            
            # Buscar producto
            producto = None
            if producto_id:
                try:
                    producto = Producto.objects.get(id=int(producto_id))
                except (Producto.DoesNotExist, ValueError, TypeError):
                    pass
            
            if not producto and nombre:
                try:
                    producto = Producto.objects.get(nombre__iexact=nombre)
                except Producto.DoesNotExist:
                    pass
                except Producto.MultipleObjectsReturned:
                    result['errors'].append(f'Fila {row_idx}: Se encontraron múltiples productos con el nombre "{nombre}". Use el ID del producto.')
                    continue
            
            if not producto:
                result['errors'].append(f'Fila {row_idx}: No se encontró el producto con ID "{producto_id}" o nombre "{nombre}".')
                continue
            
            # Parsear fecha de caducidad
            fecha_caducidad = None
            if fecha_cad:
                try:
                    if isinstance(fecha_cad, str):
                        # Intentar parsear formato DD/MM/YYYY
                        from datetime import datetime
                        fecha_caducidad = datetime.strptime(fecha_cad, '%d/%m/%Y').date()
                    else:
                        # Si es un objeto datetime de Excel
                        fecha_caducidad = fecha_cad.date() if hasattr(fecha_cad, 'date') else fecha_cad
                except Exception as e:
                    result['warnings'].append(f'Fila {row_idx}: No se pudo parsear la fecha de caducidad "{fecha_cad}". Use formato DD/MM/YYYY.')
            
            # Agregar producto procesado
            productos_procesados.append({
                'producto_id': producto.id,
                'producto_nombre': producto.nombre,
                'cantidad': int(cantidad),
                'costo_unitario': float(costo),
                'fecha_caducidad': fecha_caducidad,
                'tasa_iva_id': producto.tasa_iva_id if producto.tasa_iva else None,
                'iva_porcentaje': float(producto.tasa_iva.porcentaje) if producto.tasa_iva else 0,
            })
        
        # Validar que se hayan procesado productos
        if not productos_procesados:
            result['errors'].append('No se pudo procesar ningún producto del archivo.')
            return result
        
        # Si hay errores críticos, no retornar datos
        if result['errors']:
            result['success'] = False
            return result
        
        result['success'] = True
        result['data'] = productos_procesados
        
        logger.info(f"[EXCEL IMPORT] Se procesaron {len(productos_procesados)} productos exitosamente")
        
    except Exception as e:
        logger.error(f"[EXCEL IMPORT] Error al procesar archivo Excel: {e}", exc_info=True)
        result['errors'].append(f'Error al procesar el archivo: {str(e)}')
    
    return result
