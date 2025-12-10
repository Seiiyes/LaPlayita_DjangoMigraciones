import json
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.views.decorators.http import require_POST
from django.views.decorators.cache import never_cache
from django.contrib import messages
from datetime import date, timedelta
from django.db import models
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from django.contrib.auth.decorators import login_required
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

from .models import Producto, Lote, Categoria, MovimientoInventario, TasaIVA
from suppliers.models import Proveedor
from .forms import ProductoForm, LoteForm, CategoriaForm
from users.decorators import check_user_role

# ----------------------------------------------
# Vistas Modernas de Inventario
# ----------------------------------------------


@never_cache
@login_required
@check_user_role(allowed_roles=['Administrador', 'Vendedor'])
def alertas_stock_list(request):
    """Vista moderna de alertas de stock bajo"""
    return render(request, 'inventory/alertas_modern.html')

@never_cache
@login_required
@require_POST
@check_user_role(allowed_roles=['Administrador'])
def producto_delete(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    producto.delete()
    messages.success(request, 'Producto eliminado exitosamente.')
    return redirect('inventory:productos_list')

@never_cache
@login_required
@require_POST
@check_user_role(allowed_roles=['Administrador', 'Vendedor'])
def categoria_create(request):
    if 'application/json' in request.headers.get('Content-Type', ''):
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'JSON inválido.'}, status=400)
        
        form = CategoriaForm(data)
        if form.is_valid():
            categoria = form.save()
            return JsonResponse({'id': categoria.pk, 'nombre': categoria.nombre}, status=201)
        else:
            return JsonResponse({'errors': form.errors}, status=400)

    # Fallback para el comportamiento no-AJAX si aún es necesario
    form = CategoriaForm(request.POST)
    if form.is_valid():
        form.save()
        messages.success(request, 'Categoría creada exitosamente.')
    else:
        for field, errors in form.errors.items():
            for error in errors:
                messages.error(request, f"Error en {form.fields[field].label}: {error}")
    return redirect('inventory:productos_list')

@login_required
@check_user_role(allowed_roles=['Administrador'])
def descartar_lote(request, lote_id):
    """Vista para descartar productos de un lote específico"""
    from django.db import transaction
    from .forms import DescartarLoteForm
    
    lote = get_object_or_404(Lote.objects.select_related('producto'), pk=lote_id)
    
    if request.method == 'POST':
        form = DescartarLoteForm(request.POST, lote=lote)
        if form.is_valid():
            try:
                with transaction.atomic():
                    cantidad = form.cleaned_data['cantidad']
                    motivo = form.cleaned_data['motivo']
                    observaciones = form.cleaned_data['observaciones']
                    
                    # Actualizar cantidad disponible del lote
                    lote.cantidad_disponible -= cantidad
                    lote.save()
                    
                    # Registrar movimiento de inventario
                    MovimientoInventario.objects.create(
                        producto=lote.producto,
                        lote=lote,
                        cantidad=-cantidad,
                        tipo_movimiento='descarte',
                        descripcion=f'Descarte por {motivo} - Lote {lote.numero_lote}. {observaciones}'
                    )
                    
                    messages.success(
                        request,
                        f'✅ Se descartaron {cantidad} unidades de {lote.producto.nombre}. '
                        f'Cantidad restante: {lote.cantidad_disponible}'
                    )
                    
                    return redirect('inventory:productos_list')
            except Exception as e:
                messages.error(request, f'❌ Error al descartar productos: {str(e)}')
    else:
        form = DescartarLoteForm(lote=lote)
    
    context = {
        'form': form,
        'lote': lote,
    }
    
    return render(request, 'inventory/descartar_lote.html', context)



# ----------------------------------------------
# Gestión de Lotes
# ----------------------------------------------

@never_cache
@login_required
@check_user_role(allowed_roles=['Administrador', 'Vendedor'])
def dashboard_inventario(request):
    """Vista principal del dashboard de inventario"""
    return render(request, 'inventory/dashboard.html')



@never_cache
@login_required
@check_user_role(allowed_roles=['Administrador', 'Vendedor'])
def productos_list_modern(request):
    """Vista moderna de listado de productos con filtros"""
    return render(request, 'inventory/productos_list_modern.html')



@never_cache
@login_required
@check_user_role(allowed_roles=['Administrador', 'Vendedor'])
def producto_create_form(request):
    """Vista para mostrar y procesar el formulario de creación de producto"""
    if request.method == 'POST':
        form = ProductoForm(request.POST)
        if form.is_valid():
            producto = form.save()
            messages.success(request, f'Producto "{producto.nombre}" creado exitosamente.')
            return redirect('inventory:productos_list')
        else:
            messages.error(request, 'Error al crear el producto. Verifica los datos.')
    else:
        form = ProductoForm()
    
    categorias = Categoria.objects.all()
    context = {
        'form': form,
        'categorias': categorias,
        'titulo': 'Nuevo Producto',
        'accion': 'Crear'
    }
    return render(request, 'inventory/producto_form.html', context)


@never_cache
@login_required
@check_user_role(allowed_roles=['Administrador', 'Vendedor'])
def producto_edit_form(request, pk):
    """Vista para mostrar y procesar el formulario de edición de producto"""
    producto = get_object_or_404(Producto, pk=pk)
    
    if request.method == 'POST':
        form = ProductoForm(request.POST, instance=producto)
        if form.is_valid():
            producto = form.save()
            messages.success(request, f'Producto "{producto.nombre}" actualizado exitosamente.')
            return redirect('inventory:productos_list')
        else:
            messages.error(request, 'Error al actualizar el producto. Verifica los datos.')
    else:
        form = ProductoForm(instance=producto)
    
    categorias = Categoria.objects.all()
    context = {
        'form': form,
        'producto': producto,
        'categorias': categorias,
        'titulo': f'Editar Producto: {producto.nombre}',
        'accion': 'Actualizar'
    }
    return render(request, 'inventory/producto_form.html', context)



@never_cache
@login_required
@check_user_role(allowed_roles=['Administrador', 'Vendedor'])
def producto_detalle(request, pk):
    """Vista de detalle completo del producto con tabs"""
    producto = get_object_or_404(Producto, pk=pk)
    valor_total = producto.stock_actual * producto.costo_promedio
    
    context = {
        'producto': producto,
        'valor_total': valor_total
    }
    return render(request, 'inventory/producto_detalle.html', context)


@never_cache
@login_required
@check_user_role(allowed_roles=['Administrador', 'Vendedor'])
def kardex_producto(request, pk):
    """Vista del kardex de un producto (legacy - redirige a detalle)"""
    return redirect('inventory:producto_detalle', pk=pk)


@never_cache
@login_required
@check_user_role(allowed_roles=['Administrador', 'Vendedor'])
def reabastecimientos_list(request):
    """Vista moderna de reabastecimientos"""
    return render(request, 'inventory/reabastecimientos_list.html')


@never_cache
@login_required
@check_user_role(allowed_roles=['Administrador'])
def ajustes_list(request):
    """Vista moderna de ajustes de inventario"""
    return render(request, 'inventory/ajustes_list.html')


@never_cache
@login_required
@check_user_role(allowed_roles=['Administrador', 'Vendedor'])
def reportes(request):
    """Vista de reportes de inventario"""
    return render(request, 'inventory/reportes.html')

@login_required
@check_user_role(allowed_roles=['Administrador', 'Vendedor'])
def exportar_productos_excel(request):
    """Exportar productos a Excel"""
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Encabezados
    headers = [
        'ID', 'Nombre', 'Descripción', 'Categoría', 'Código de Barras',
        'Stock Actual', 'Stock Mínimo', 'Precio Unitario', 'Costo Promedio',
        'Valor Inventario', 'Estado Stock', 'Estado', 'Fecha Creación'
    ]
    
    # Escribir encabezados
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Obtener productos
    productos = Producto.objects.select_related('categoria').filter(estado='activo')
    
    # Aplicar filtros si existen
    search = request.GET.get('search')
    if search:
        productos = productos.filter(nombre__icontains=search)
    
    categoria = request.GET.get('categoria')
    if categoria:
        productos = productos.filter(categoria_id=categoria)
    
    # Escribir datos
    for row, producto in enumerate(productos, 2):
        # Calcular estado de stock
        if producto.stock_actual <= 0:
            estado_stock = 'SIN_STOCK'
        elif producto.stock_actual <= producto.stock_critico:
            estado_stock = 'STOCK_CRITICO'
        elif producto.stock_actual <= producto.stock_minimo:
            estado_stock = 'STOCK_BAJO'
        else:
            estado_stock = 'NORMAL'
        
        # Calcular valor inventario
        valor_inventario = producto.stock_actual * producto.costo_promedio
        
        data = [
            producto.id,
            producto.nombre,
            producto.descripcion or '',
            producto.categoria.nombre if producto.categoria else '',
            producto.codigo_barras or '',
            producto.stock_actual,
            producto.stock_minimo,
            float(producto.precio_unitario),
            float(producto.costo_promedio),
            float(valor_inventario),
            estado_stock,
            producto.estado,
            producto.fecha_creacion.strftime('%Y-%m-%d %H:%M:%S') if producto.fecha_creacion else ''
        ]
        
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    filename = f'productos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Guardar workbook
    wb.save(response)
    
    return response