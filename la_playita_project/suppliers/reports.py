"""
Generación de reportes para reabastecimientos en PDF y Excel
"""
from io import BytesIO
from django.http import HttpResponse
from django.conf import settings
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
from datetime import datetime
import os


def generate_reabastecimiento_pdf(reabastecimiento):
    """
    Genera un PDF elegante para un reabastecimiento
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    # Contenedor de elementos
    elements = []
    styles = getSampleStyleSheet()
    
    # Estilos personalizados
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#0d6efd'),
        spaceAfter=30,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=12,
        textColor=colors.HexColor('#6c757d'),
        spaceAfter=20,
        alignment=TA_CENTER
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#212529'),
        spaceAfter=12,
        spaceBefore=20,
        fontName='Helvetica-Bold'
    )
    
    # Logo (si existe)
    try:
        if settings.STATIC_ROOT and os.path.exists(settings.STATIC_ROOT):
            logo_path = os.path.join(settings.STATIC_ROOT, 'core/img/la-playita-logo.png')
        else:
            logo_path = os.path.join(settings.BASE_DIR, 'core/static/core/img/la-playita-logo.png')
        
        if os.path.exists(logo_path):
            logo = Image(logo_path, width=1.5*inch, height=1.5*inch)
            logo.hAlign = 'CENTER'
            elements.append(logo)
            elements.append(Spacer(1, 0.3*inch))
    except Exception as e:
        # Si hay error con el logo, continuar sin él
        pass
    
    # Título
    elements.append(Paragraph(f"Orden de Reabastecimiento #{reabastecimiento.id}", title_style))
    elements.append(Paragraph(f"La Playita - Sistema de Gestión", subtitle_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Información general
    info_data = [
        ['Proveedor:', reabastecimiento.proveedor.nombre_empresa],
        ['Fecha de Solicitud:', reabastecimiento.fecha.strftime('%d/%m/%Y %H:%M')],
        ['Estado:', reabastecimiento.get_estado_display()],
        ['Forma de Pago:', reabastecimiento.get_forma_pago_display()],
    ]
    
    if reabastecimiento.observaciones:
        info_data.append(['Observaciones:', reabastecimiento.observaciones])
    
    info_table = Table(info_data, colWidths=[2*inch, 4.5*inch])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f8f9fa')),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#495057')),
        ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#212529')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
    ]))
    
    elements.append(info_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Productos
    elements.append(Paragraph("Productos Solicitados", heading_style))
    
    # Tabla de productos
    productos_data = [['Producto', 'Cantidad', 'Costo Unit.', 'IVA', 'Subtotal', 'Total']]
    
    for detalle in reabastecimiento.reabastecimientodetalle_set.all():
        subtotal = float(detalle.cantidad * detalle.costo_unitario)
        iva = float(detalle.iva or 0)
        total = subtotal + iva
        
        productos_data.append([
            detalle.producto.nombre,
            str(detalle.cantidad),
            f"${detalle.costo_unitario:,.0f}",
            f"${iva:,.0f}",
            f"${subtotal:,.0f}",
            f"${total:,.0f}"
        ])
    
    productos_table = Table(productos_data, colWidths=[2.5*inch, 0.8*inch, 1*inch, 0.9*inch, 1*inch, 1*inch])
    productos_table.setStyle(TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d6efd')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        
        # Body
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#212529')),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),
        ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        
        # Grid
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
    ]))
    
    elements.append(productos_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Totales
    totales_data = [
        ['Subtotal:', f"${float(reabastecimiento.costo_total):,.0f}"],
        ['IVA:', f"${float(reabastecimiento.iva):,.0f}"],
        ['Total:', f"${float(reabastecimiento.costo_total + reabastecimiento.iva):,.0f}"],
    ]
    
    totales_table = Table(totales_data, colWidths=[5.5*inch, 1.5*inch])
    totales_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('TEXTCOLOR', (0, 0), (-1, 1), colors.HexColor('#495057')),
        ('TEXTCOLOR', (0, 2), (-1, 2), colors.HexColor('#198754')),
        ('FONTSIZE', (0, 2), (-1, 2), 14),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('LINEABOVE', (0, 2), (-1, 2), 2, colors.HexColor('#198754')),
    ]))
    
    elements.append(totales_table)
    elements.append(Spacer(1, 0.5*inch))
    
    # Footer
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.HexColor('#6c757d'),
        alignment=TA_CENTER
    )
    
    elements.append(Paragraph(
        f"Documento generado el {datetime.now().strftime('%d/%m/%Y a las %H:%M')}",
        footer_style
    ))
    elements.append(Paragraph("La Playita - Sistema de Gestión de Inventario", footer_style))
    
    # Construir PDF
    doc.build(elements)
    
    buffer.seek(0)
    return buffer


def generate_reabastecimiento_excel(reabastecimiento):
    """
    Genera un archivo Excel elegante para un reabastecimiento
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"Orden {reabastecimiento.id}"
    
    # Estilos
    header_fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=16, color="0D6EFD")
    subtitle_font = Font(size=11, color="6C757D")
    bold_font = Font(bold=True, size=11)
    total_font = Font(bold=True, size=14, color="198754")
    
    border = Border(
        left=Side(style='thin', color='DEE2E6'),
        right=Side(style='thin', color='DEE2E6'),
        top=Side(style='thin', color='DEE2E6'),
        bottom=Side(style='thin', color='DEE2E6')
    )
    
    # Logo (si existe)
    try:
        if settings.STATIC_ROOT and os.path.exists(settings.STATIC_ROOT):
            logo_path = os.path.join(settings.STATIC_ROOT, 'core/img/la-playita-logo.png')
        else:
            logo_path = os.path.join(settings.BASE_DIR, 'core/static/core/img/la-playita-logo.png')
        
        if os.path.exists(logo_path):
            img = XLImage(logo_path)
            img.width = 100
            img.height = 100
            ws.add_image(img, 'A1')
    except Exception as e:
        # Si hay error con el logo, continuar sin él
        pass
    
    # Título
    ws.merge_cells('A1:F1')
    ws['A1'] = f"Orden de Reabastecimiento #{reabastecimiento.id}"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells('A2:F2')
    ws['A2'] = "La Playita - Sistema de Gestión"
    ws['A2'].font = subtitle_font
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Información general
    row = 4
    info_data = [
        ('Proveedor:', reabastecimiento.proveedor.nombre_empresa),
        ('Fecha de Solicitud:', reabastecimiento.fecha.strftime('%d/%m/%Y %H:%M')),
        ('Estado:', reabastecimiento.get_estado_display()),
        ('Forma de Pago:', reabastecimiento.get_forma_pago_display()),
    ]
    
    if reabastecimiento.observaciones:
        info_data.append(('Observaciones:', reabastecimiento.observaciones))
    
    for label, value in info_data:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = bold_font
        ws[f'A{row}'].fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        ws[f'B{row}'] = value
        ws.merge_cells(f'B{row}:F{row}')
        row += 1
    
    row += 2
    
    # Tabla de productos
    ws[f'A{row}'] = "Productos Solicitados"
    ws[f'A{row}'].font = Font(bold=True, size=14)
    ws.merge_cells(f'A{row}:F{row}')
    row += 1
    
    # Headers
    headers = ['Producto', 'Cantidad', 'Costo Unitario', 'IVA', 'Subtotal', 'Total']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    row += 1
    
    # Datos de productos
    for detalle in reabastecimiento.reabastecimientodetalle_set.all():
        subtotal = float(detalle.cantidad * detalle.costo_unitario)
        iva = float(detalle.iva or 0)
        total = subtotal + iva
        
        ws.cell(row=row, column=1, value=detalle.producto.nombre).border = border
        ws.cell(row=row, column=2, value=detalle.cantidad).border = border
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='right')
        ws.cell(row=row, column=3, value=float(detalle.costo_unitario)).border = border
        ws.cell(row=row, column=3).number_format = '$#,##0'
        ws.cell(row=row, column=3).alignment = Alignment(horizontal='right')
        ws.cell(row=row, column=4, value=iva).border = border
        ws.cell(row=row, column=4).number_format = '$#,##0'
        ws.cell(row=row, column=4).alignment = Alignment(horizontal='right')
        ws.cell(row=row, column=5, value=subtotal).border = border
        ws.cell(row=row, column=5).number_format = '$#,##0'
        ws.cell(row=row, column=5).alignment = Alignment(horizontal='right')
        ws.cell(row=row, column=6, value=total).border = border
        ws.cell(row=row, column=6).number_format = '$#,##0'
        ws.cell(row=row, column=6).alignment = Alignment(horizontal='right')
        
        # Alternar colores de fila
        if row % 2 == 0:
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        
        row += 1
    
    row += 2
    
    # Totales
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "Subtotal:"
    ws[f'A{row}'].font = bold_font
    ws[f'A{row}'].alignment = Alignment(horizontal='right')
    ws[f'F{row}'] = float(reabastecimiento.costo_total)
    ws[f'F{row}'].number_format = '$#,##0'
    ws[f'F{row}'].font = bold_font
    ws[f'F{row}'].alignment = Alignment(horizontal='right')
    
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "IVA:"
    ws[f'A{row}'].font = bold_font
    ws[f'A{row}'].alignment = Alignment(horizontal='right')
    ws[f'F{row}'] = float(reabastecimiento.iva)
    ws[f'F{row}'].number_format = '$#,##0'
    ws[f'F{row}'].font = bold_font
    ws[f'F{row}'].alignment = Alignment(horizontal='right')
    
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "Total:"
    ws[f'A{row}'].font = total_font
    ws[f'A{row}'].alignment = Alignment(horizontal='right')
    ws[f'F{row}'] = float(reabastecimiento.costo_total + reabastecimiento.iva)
    ws[f'F{row}'].number_format = '$#,##0'
    ws[f'F{row}'].font = total_font
    ws[f'F{row}'].alignment = Alignment(horizontal='right')
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    
    # Guardar en buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer
